# -*- coding: utf-8 -*-
"""
Spyder Editor
Created on Mon June  20 15:41:12 2018

@author: sintu

"""
# Importing the libraries
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#dataset
x = pd.read_excel('features1.xlsx')
y = pd.read_excel('turker_scores1.xlsx')
yt = y.iloc[0:,0:1]

#Creating Mutual Information file
import openpyxl
from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet(title='Mutual Information')
ws1 = openpyxl.load_workbook('features.xlsx')
ws2 = openpyxl.load_workbook('turker_scores.xlsx')
sheet2 = ws1.get_sheet_by_name('Feature')
sheet3 = ws2.get_sheet_by_name('turker_scores')

for i in range(1,23):
    ws.cell(row=1,column=i+1,value=(sheet2.cell(row=1,column=i+1).value))
    ws.cell(row=i+1,column=1,value=(sheet3.cell(row=1,column=i+1).value))
wb.save('MI_score.xlsx')
      
#Calculating Mutual Information
from sklearn.feature_selection import mutual_info_regression
for i in range(1,20):
    mi = []
    y1 = []
    y1 = y.iloc[0:,i-1:i]
    mi = mutual_info_regression(x,y1)
    for j in range(1,22):
        ws.cell(row=i+1,column=j+1,value=mi[j-1])

