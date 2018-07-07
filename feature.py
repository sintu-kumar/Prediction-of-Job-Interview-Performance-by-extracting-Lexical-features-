# -*- coding: utf-8 -*-
"""
Created on Tue Jun  5 16:47:15 2018

@author: sintu
"""

import openpyxl
import numpy as np
import pandas as pd
import nltk

ws2 = openpyxl.load_workbook('interview_transcripts_by_turker_sheet.xlsx')
list = ['ah', 'Ah', 'Ahh','ahh','uhm', 'Uhm', 'Uhmmm', 'um','Um', 'uh', 'Uhh', 'Umm', 'ummm', 'Mmhmm', 'Uhhhh', 'Hmm', 'uhhhh', 'uhh', 'umm', 'Uh']
posEmotion = ['hope', 'improve', 'kind', 'love', 'happy', 'pretty', 'good']
negEmotion = ['bad', 'fool', 'hate', 'lose', 'hate', 'worthless', 'enemy']
Anxiety = ['nervous', 'obsessed', 'panic', 'shy', 'afraid']
Anger = ['agitate', 'bother', 'confront', 'disgust', 'hate', 'kill', 'pissed']
Sadness = ['fail', 'grief', 'hurt', 'inferior', 'cry', 'sad']  
Cognitive = ['cause', 'know', 'ought', 'learn', 'make', 'notice']
Inhibition = ['refrain', 'prohibit', 'prevent', 'stop', 'block', 'constrain']
Perceptual = ['observe', 'experience', 'view', 'watch']
Work = ['project', 'study', 'thesis', 'class', 'work', 'university']
Articles = ['a', 'an', 'the']
Negations = ['no', 'never', 'none', 'cannot', 'don’t']
Quantiﬁers = ['all', 'best', 'bunch', 'few', 'ton', 'unique']

tentative_language = ['maybe', 'perhaps', 'guess']
Confidence = ['completed', 'finished', 'assured']

sheet1 = []
total1 = []
total2 = []
total3 = []
total4 = []
total44 = []
total5 = []
total6 = []
total7 = []
total8 = []
total9 = []
total10 = []
total11 = []
total12 = []
total13 = []
total = []
p_o_s = []

verb_count = []
adverb_count = []
prep_count = []
conj_count = []
Icount = []
WEcount = []
THEYcount = []
sheet1 = ws2.get_sheet_names()
for i in range(0,138):
    sheetx = ws2.get_sheet_by_name(sheet1[i])
    words = []
    mx = 0
    mx1 = 0
    mx2 = 0
    mx3 = 0
    mx4 = 0
    mx44 = 0
    mx5 = 0
    mx6 = 0
    mx7 = 0
    mx8 = 0
    mx9 = 0
    mx10 = 0
    mx11 = 0
    mx12 = 0
    mx13 = 0
    verb = []
    adverb = []
    preposition = []
    conjuction = []
    x = sheetx.max_row
    for j in range(2,x+1,2):
        words = words + ((sheetx.cell(row=j, column=2)).value).split()
    for k in range(0,20):
        mx = mx + words.count(list[k])
    total.append(mx)
    Icount.append(words.count('I'))
    WEcount.append(words.count('we'))
    THEYcount.append(words.count('they'))
    
    for word,pos in nltk.pos_tag(words):
        if(pos=='VB' or pos=='VBD' or pos=='VBG' or pos=='VBN' or pos=='VBP' or pos=='VBZ'):
            verb.append(word)
        if(pos=='RB' or pos=='RBR' or pos=='RBS'):
            adverb.append(word)
        if(pos=='IN'):
            preposition.append(word)
        if(pos=='CC'):
            conjuction.append(word)
            
    verb_count.append(len(verb))
    adverb_count.append(len(adverb))
    prep_count.append(len(preposition))
    conj_count.append(len(conjuction))
    
    for p in range(0,7):
        mx1 = mx1 + words.count(posEmotion[p])
    total1.append(mx1)
    for p in range(0,7):
        mx2 = mx2 + words.count(negEmotion[p])
    total2.append(mx2)
    for p in range(0,5):
        mx3 = mx3 + words.count(Anxiety[p])
    total3.append(mx3)
    for p in range(0,7):
        mx4 = mx4 + words.count(Anger[p])
    total4.append(mx4)
    for p in range(0,6):
        mx44 = mx44 + words.count(Sadness[p])
    total44.append(mx44)
    for p in range(0,6):
        mx5 = mx5 + words.count(Cognitive[p])
    total5.append(mx5)
    for p in range(0,6):
        mx6 = mx6 + words.count(Inhibition[p])
    total6.append(mx6)
    for p in range(0,4):
        mx7 = mx7 + words.count(Perceptual[p])
    total7.append(mx7)
    for p in range(0,6):
        mx8 = mx8 + words.count(Work[p])
    total8.append(mx8)
    for p in range(0,3):
        mx9 = mx9 + words.count(Articles[p])
    total9.append(mx9)
    for p in range(0,5):
        mx10 = mx10 + words.count(Negations[p])
    total10.append(mx10)
    for p in range(0,3):
        mx12 = mx12 + words.count(tentative_language[p])
    total12.append(mx12)
    for p in range(0,3):
        mx13 = mx13 + words.count(Confidence[p])
    total13.append(mx13)
    
    
    
           
        
from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet(title='Feature')
ws.append(['People','I','We','they','PosEmotion','NegEmotion','anxiety','Anger','Sadness','Cognitive','Inhibition','Perceptual','work','Article','Verb','Adverb','Preposition','Conjuction','Negations','Quqntifier','Tentative_Lang.','Confidence','FillerWords'])
import openpyxl
ws1 = openpyxl.load_workbook('interview_transcripts_by_turker_sheet.xlsx')
sheet2 = []
sheet2 = ws1.get_sheet_names()
for i in range(0,138):
    ws.cell(row=i+2,column=1,value=sheet2[i])

for j in range(0,138):
    ws.cell(row=j+2,column=23,value=total[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=2,value=Icount[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=15,value=verb_count[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=16,value=adverb_count[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=17,value=prep_count[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=18,value=conj_count[j])

for j in range(0,138):
    ws.cell(row=j+2,column=3,value=WEcount[j])

for j in range(0,138):
    ws.cell(row=j+2,column=4,value=THEYcount[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=5,value=total1[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=6, value=total2[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=7,value=total3[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=8,value=total4[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=9,value=total44[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=10,value=total5[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=11,value=total6[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=12,value=total7[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=13,value=total8[j])

for j in range(0,138):
    ws.cell(row=j+2,column=14,value=total9[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=19,value=total10[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=21,value=total12[j])
    
for j in range(0,138):
    ws.cell(row=j+2,column=22,value=total13[j])

ws.sheet_properties.tabColor = "1072BA" 
wb.save('features.xlsx')

        
        
        
        


